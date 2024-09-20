# automatizar envio de mensagens p 20k pessoass
from openpyxl import load_workbook
from urllib.parse import quote
import webbrowser
from time import sleep
import pyautogui

soma_contatos_enviados = 0 # soma total de contatos enviados
# descrever os passos manuais e dps transformar isso em código
webbrowser.open('https://web.whatsapp.com/')
sleep(20)

# ler planlha e guardar informações sobre, nome, numero
arquivo = load_workbook('clientes.xlsx')
pagina_clientes = arquivo['Planilha1']

for linha in pagina_clientes.iter_rows(min_row=2): # iter_rows, percorre por todas as linhas da planilha, iniando da segunda linha
    nome = linha[0].value # pegando o valor/nome da primeira posisão da linha
    numero = linha[1].value # .value retorna valor

    if nome is None or numero is None: # caso tenha um campo no excel vazio2️
        print(f"Dados inválidos: {nome}, {numero}")
        continue

    mensagem = f'Fala {nome}, aqui é o Jakson Capixaba. Estou na luta. Me ajude a multiplicar e me fazer vearador de Natal. O vereador da torcida do ABC. Vote ✅2️⃣0️⃣.2️⃣3️⃣4️⃣✅, a bola da vez.'


# criar link personalizados do whatsapp e enviar mensagens para cada cliente, com base na planilha  
    try:
        # mensagem
        link_mensagem_whatsapp = f'https://web.whatsapp.com/send?phone={numero}&text={quote(mensagem)}'
        webbrowser.open(link_mensagem_whatsapp)
        sleep(8)
        pyautogui.click(x=1851, y=964)
        sleep(3)

        # imagem
        pyautogui.click(x=765, y=965)
        sleep(2)
        pyautogui.click(x=854, y=710)
        sleep(2)
        pyautogui.click(x=548, y=207)
        sleep(2)
        pyautogui.press('enter')
        sleep(2)
        pyautogui.press('enter')
        sleep(1)

        # audio
        sleep(1)
        pyautogui.moveTo(-1925, 264) # pega o audio
        pyautogui.dragTo(996, 713, duration=2) # e puxa ate o destino hahah
        sleep(3) # tempo p carregar 
        pyautogui.press('enter')
        sleep(2)
        pyautogui.hotkey('ctrl', 'w')
        soma_contatos_enviados += 1
        print(soma_contatos_enviados)
    except Exception as e:
        print(f'Não foi possível abrir a conversa com {nome}. Erro: {str(e)}')
        with open('erros.csv', 'a', newline='', encoding='utf-8') as arquivo:
            arquivo.write(f'{nome}, {numero}\n')